import streamlit as st
import pandas as pd
import numpy as np
from urllib.parse import urlparse
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import zipfile
from datetime import datetime
import math
import plotly.express as px
from io import BytesIO

def main():
    st.set_page_config(
        page_title="Audit Sémantique - Gap Content SEO",
        page_icon="🔍",
        layout="wide"
    )
    
    st.title("Outil d'Audit Sémantique - Gap Content SEO")
    st.markdown("---")
    
    # Présentation de l'outil
    with st.expander("ℹ️ En quoi consiste ce script ?"):
        st.markdown("""
        Ce script sert à identifier rapidement les opportunités de trafic sur base du gap content (contenu non adressé) entre votre domaine et un ou plusieurs domaines concurrents.
        
        **Comment faire ?**
        
        **1. Préparez vos fichiers**
        * Exportez vos données SEO (Semrush, Ahrefs, etc.) en CSV/Excel,
        * **Important** : Téléchargez votre site en premier si vous voulez aller plus vite.
        
        **2. Configurez l'analyse**
        * **Source** : Semrush/Ahrefs/Custom,
        * **Concurrents minimum** : 2 (recommandé),
        * **Position max** : Top 10 (page 1),
        * **Filtres** : Volume min selon vos besoins.
        """)
    
    st.markdown("---")
    
    # Sidebar pour la configuration
    with st.sidebar:
        st.header("⚙️ Configuration")
        
        # Section 1: Type de source de données
        st.subheader("1. Source des données")
        data_source = st.selectbox(
            "Type de fichier d'export",
            ["Semrush", "Ahrefs", "Custom"]
        )
        
        # Section 2: Configuration des colonnes (si Custom)
        if data_source == "Custom":
            st.subheader("2. Mapping des colonnes")
            col_keyword = st.text_input("Nom colonne Mot-clé", "Keyword")
            col_domain = st.text_input("Nom colonne Domaine", "Domain")
            col_position = st.text_input("Nom colonne Position", "Position")
            col_volume = st.text_input("Nom colonne Volume de recherche", "Search Volume")
            col_difficulty = st.text_input("Nom colonne Difficulté", "Keyword Difficulty")
            col_intent = st.text_input("Nom colonne Intention", "Keyword Intents")
            col_url = st.text_input("Nom colonne URL", "URL")
        else:
            # Configuration prédéfinie pour Semrush/Ahrefs
            if data_source == "Semrush":
                col_mapping = {
                    'keyword': 'Keyword',
                    'domain': 'URL',  # On extraira le domaine de l'URL
                    'position': 'Position',
                    'volume': 'Search Volume',
                    'difficulty': 'Keyword Difficulty',
                    'intent': 'Keyword Intents',
                    'url': 'URL'
                }
            else:  # Ahrefs
                col_mapping = {
                    'keyword': 'Keyword',
                    'domain': None,  # Pas de colonne URL dans ce format, on utilisera le nom de fichier
                    'position': 'Average position',
                    'volume': 'Volume',
                    'difficulty': None,  # Pas de difficulté dans ce format
                    'intent': None,  # Pas d'intention dans ce format
                    'url': None,  # Pas d'URL spécifique
                    'traffic': 'Organic traffic'
                }
        
        # Section 3: Critères de filtrage
        st.subheader("2. Critères Gap Content" if data_source != "Custom" else "3. Critères Gap Content")
        min_competitors = st.selectbox(
            "Nombre minimum de concurrents positionnés",
            [1, 2, 3]
        )
        
        # Position maximum avec curseur et saisie manuelle
        st.write("Position maximum")
        max_position = st.slider(
            "",
            min_value=1,
            max_value=100,
            value=10,
            step=1,
            label_visibility="collapsed"
        )
        
        max_position = st.number_input(
            "Saisissez manuellement",
            min_value=1,
            max_value=100,
            value=max_position,
            step=1
        )
        
        # Filtres supplémentaires
        st.subheader("3. Filtres supplémentaires" if data_source != "Custom" else "4. Filtres supplémentaires")
        min_volume = st.number_input("Volume de recherche minimum", min_value=0, value=0)
        max_difficulty = st.number_input("Difficulté maximum", min_value=0, max_value=100, value=100)
        
        # Filtre termes de marque
        brand_terms = st.text_input(
            "Termes de marque (séparés par des virgules)",
            placeholder="marque1, marque2, marque3",
            help="Mots-clés contenant ces termes seront filtrés de l'analyse"
        )

    # Zone principale
    st.header("📁 Import des fichiers")
    
    # Upload des fichiers
    uploaded_files = st.file_uploader(
        "Téléchargez vos fichiers CSV/Excel (le premier sera considéré comme votre domaine principal)",
        accept_multiple_files=True,
        type=['csv', 'xlsx', 'xls']
    )
    
    # Identification du domaine principal
    if uploaded_files:
        st.subheader("Identification du domaine principal")
        
        # Option 1: Premier fichier par défaut
        main_domain_option = st.radio(
            "Comment identifier votre domaine principal ?",
            ["Premier fichier téléchargé", "Sélection manuelle"]
        )
        
        if main_domain_option == "Sélection manuelle":
            main_domain = st.text_input("Saisissez votre domaine principal (ex: www.monsite.com)")
        else:
            main_domain = None
    
    # Bouton d'analyse
    if uploaded_files and st.button("Lancer l'analyse", type="primary"):
        with st.spinner("Analyse en cours..."):
            try:
                # Traitement des fichiers
                all_data = process_files(uploaded_files, data_source, locals())
                
                if all_data is not None and not all_data.empty:
                    # Identification du domaine principal
                    if main_domain_option == "Premier fichier téléchargé":
                        first_file_data = load_file(uploaded_files[0], data_source, locals())
                        if not first_file_data.empty:
                            main_domain = extract_domain_from_data(first_file_data)
                            st.success(f"Domaine principal détecté automatiquement : {main_domain}")
                    
                    # Filtrage des termes de marque
                    if brand_terms.strip():
                        all_data = filter_brand_terms(all_data, brand_terms)
                    
                    # Analyse du gap content
                    gap_analysis = perform_gap_analysis(
                        all_data, 
                        main_domain, 
                        min_competitors, 
                        max_position,
                        min_volume,
                        max_difficulty
                    )
                    
                    # Analyse du domaine principal
                    main_domain_analysis = analyze_main_domain(all_data, main_domain) if main_domain else None
                    
                    if gap_analysis['gap_content'].empty:
                        st.warning("Aucune opportunité de gap content trouvée avec ces critères.")
                    else:
                        # Génération du fichier Excel
                        excel_file = generate_excel_report(gap_analysis, main_domain, main_domain_analysis)
                        
                        # Affichage d'un résumé simple
                        st.success(f"Analyse terminée ! {len(gap_analysis['gap_content'])} opportunités trouvées.")
                        
                        if not gap_analysis['gap_content'].empty:
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.metric("Total opportunités", len(gap_analysis['gap_content']))
                            with col2:
                                avg_volume = gap_analysis['gap_content']['volume'].mean()
                                st.metric("Volume moyen", f"{avg_volume:,.0f}")
                            with col3:
                                total_volume = gap_analysis['gap_content']['volume'].sum()
                                st.metric("Volume total", f"{total_volume:,.0f}")
                        
                        # Bouton de téléchargement
                        st.download_button(
                            label="Télécharger le rapport Excel",
                            data=excel_file,
                            file_name=f"gap_content_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                else:
                    st.error("Impossible de traiter les fichiers téléchargés.")
                    
            except Exception as e:
                st.error(f"Erreur lors de l'analyse : {str(e)}")


def extract_domain(url):
    """Extrait le domaine racine d'une URL (sans sous-domaines)"""
    try:
        if pd.isna(url) or url == '':
            return ''
        if not url.startswith(('http://', 'https://')):
            url = 'https://' + url
        parsed = urlparse(url)
        domain = parsed.netloc.lower()
        
        # Extraire le domaine racine (enlever les sous-domaines)
        domain_parts = domain.split('.')
        if len(domain_parts) >= 2:
            # Garder seulement les deux dernières parties (domaine.tld)
            # Gérer les cas spéciaux comme .co.uk, .com.fr, etc.
            if len(domain_parts) >= 3 and domain_parts[-2] in ['co', 'com', 'net', 'org', 'gov']:
                root_domain = '.'.join(domain_parts[-3:])
            else:
                root_domain = '.'.join(domain_parts[-2:])
            return root_domain
        return domain
    except:
        return ''


def extract_domain_from_data(df):
    """Extrait le domaine racine principal du premier fichier"""
    if 'domain' in df.columns and not df['domain'].empty:
        # Prendre le domaine le plus fréquent
        return df['domain'].value_counts().index[0]
    return None


def load_file(file, data_source, config):
    """Charge un fichier CSV ou Excel et normalise les colonnes"""
    try:
        if file.name.endswith('.csv'):
            # Détection automatique de l'encodage pour les CSV
            df = None
            encodings_to_try = ['utf-8', 'utf-16', 'utf-8-sig', 'latin-1', 'cp1252', 'utf-16-le', 'utf-16-be']
            separators_to_try = [',', '\t', ';']
            
            for encoding in encodings_to_try:
                for separator in separators_to_try:
                    try:
                        # Reset du pointeur de fichier
                        file.seek(0)
                        df = pd.read_csv(file, encoding=encoding, sep=separator)
                        
                        # Vérifier si le DataFrame a du sens (plus d'une colonne)
                        if len(df.columns) > 1:
                            st.success(f"Fichier {file.name} chargé avec l'encodage {encoding} et séparateur '{separator}'")
                            break
                    except (UnicodeDecodeError, UnicodeError, pd.errors.ParserError):
                        continue
                    except Exception as e:
                        continue
                if df is not None and len(df.columns) > 1:
                    break
            
            if df is None or len(df.columns) <= 1:
                # Tentative avec détection automatique Python
                try:
                    file.seek(0)
                    # Essayer avec engine python qui est plus permissif
                    df = pd.read_csv(file, encoding='utf-8', sep=None, engine='python', error_bad_lines=False)
                    st.warning(f"Fichier {file.name} chargé avec détection automatique")
                except:
                    raise ValueError(f"Impossible de détecter l'encodage du fichier {file.name}")
        else:
            # Pour les fichiers Excel, pandas gère automatiquement l'encodage
            df = pd.read_excel(file)
        
        # Debug : afficher les colonnes disponibles
        st.info(f"Colonnes détectées dans {file.name}: {list(df.columns)}")
        
        # Normalisation des colonnes selon la source
        if data_source == "Semrush":
            if 'Keyword' in df.columns:
                df['domain'] = df['URL'].apply(extract_domain)
                df = df.rename(columns={
                    'Keyword': 'keyword',
                    'Position': 'position',
                    'Search Volume': 'volume',
                    'Keyword Difficulty': 'difficulty',
                    'Keyword Intents': 'intent',
                    'URL': 'url'
                })
        elif data_source == "Ahrefs":
            # Extraction du domaine depuis le nom du fichier pour Ahrefs
            filename = file.name
            domain_from_filename = 'unknown.com'
            
            # Extraction améliorée du domaine depuis le nom de fichier
            if '-organic' in filename:
                domain_part = filename.split('-organic')[0]  # Prendre la partie avant "-organic"
                
                # Nettoyer le domaine extrait
                if domain_part.startswith('www.'):
                    domain_part = domain_part[4:]  # Enlever "www."
                
                # Gérer les cas spéciaux comme "invivo-group.com-fr"
                if domain_part.endswith('-fr'):
                    domain_part = domain_part[:-3]  # Enlever "-fr"
                elif domain_part.endswith('.com-fr'):
                    domain_part = domain_part.replace('.com-fr', '.com')
                elif domain_part.endswith('.fr-fr'):
                    domain_part = domain_part.replace('.fr-fr', '.fr')
                
                domain_from_filename = domain_part
            elif '.' in filename:
                # Fallback : prendre la partie avant le premier tiret ou espace
                domain_part = filename.split('-')[0].split('_')[0].split(' ')[0]
                if domain_part.startswith('www.'):
                    domain_part = domain_part[4:]
                domain_from_filename = domain_part
            
            # S'assurer que le domaine a une extension valide
            if not ('.' in domain_from_filename and any(ext in domain_from_filename for ext in ['.com', '.fr', '.org', '.net', '.coop'])):
                domain_from_filename = domain_from_filename + '.com'
            
            df['domain'] = extract_domain(f"https://{domain_from_filename}")
            df['url'] = f"https://{domain_from_filename}"  # URL générique
            
            # Debug : afficher le domaine extrait
            st.info(f"Domaine extrait pour {filename}: {domain_from_filename} → {df['domain'].iloc[0] if len(df) > 0 else 'N/A'}")
            
            # Mapping flexible des colonnes Ahrefs (CSV vs Excel peuvent différer)
            column_mapping = {}
            
            # Chercher les colonnes par nom (insensible à la casse)
            for col in df.columns:
                col_lower = col.lower().strip()
                if 'keyword' in col_lower:
                    column_mapping['keyword'] = col
                elif 'volume' in col_lower and 'traffic' not in col_lower and 'location' not in col_lower:
                    column_mapping['volume'] = col
                elif ('position' in col_lower or 'rank' in col_lower) and 'average' in col_lower:
                    column_mapping['position'] = col
                elif 'traffic' in col_lower and 'organic' in col_lower:
                    column_mapping['traffic'] = col
                elif 'difficulty' in col_lower or 'kd' in col_lower:
                    column_mapping['difficulty'] = col
            
            # Debug : afficher le mapping trouvé
            st.info(f"Mapping des colonnes pour {filename}: {column_mapping}")
            
            # Renommer les colonnes trouvées
            if column_mapping:
                df = df.rename(columns=column_mapping)
            
            # S'assurer que les colonnes essentielles existent
            if 'keyword' not in df.columns:
                # Essayer de trouver une colonne qui pourrait être le mot-clé
                possible_keyword_cols = [col for col in df.columns if any(term in col.lower() for term in ['keyword', 'query', 'term'])]
                if possible_keyword_cols:
                    df = df.rename(columns={possible_keyword_cols[0]: 'keyword'})
                else:
                    raise ValueError(f"Impossible de trouver une colonne 'keyword' dans {file.name}")
            
            # Vérifier si la colonne volume existe et est correcte
            if 'volume' not in df.columns:
                # Chercher d'autres colonnes possibles pour le volume
                possible_volume_cols = [col for col in df.columns if 'volume' in col.lower() and 'location' not in col.lower()]
                if possible_volume_cols:
                    df = df.rename(columns={possible_volume_cols[0]: 'volume'})
                    st.info(f"Colonne volume trouvée: {possible_volume_cols[0]} → volume")
                else:
                    df['volume'] = 0  # Valeur par défaut
                    st.warning(f"Aucune colonne volume trouvée dans {filename}, utilisation de 0 par défaut")
            
            # Ajouter des colonnes manquantes avec des valeurs par défaut
            if 'difficulty' not in df.columns:
                df['difficulty'] = 50  # Valeur par défaut
            if 'intent' not in df.columns:
                df['intent'] = 'unknown'  # Valeur par défaut
            if 'position' not in df.columns:
                df['position'] = 50  # Valeur par défaut
                st.warning(f"Aucune colonne position trouvée dans {filename}, utilisation de 50 par défaut")
                
        else:  # Custom
            df['domain'] = df[config.get('col_domain', 'URL')].apply(extract_domain)
            df = df.rename(columns={
                config.get('col_keyword', 'Keyword'): 'keyword',
                config.get('col_position', 'Position'): 'position',
                config.get('col_volume', 'Search Volume'): 'volume',
                config.get('col_difficulty', 'Keyword Difficulty'): 'difficulty',
                config.get('col_intent', 'Keyword Intents'): 'intent',
                config.get('col_url', 'URL'): 'url'
            })
        
        # Vérification finale des colonnes essentielles
        required_columns = ['keyword', 'domain']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Colonnes manquantes après traitement: {missing_columns}")
        
        # Nettoyage des données
        initial_count = len(df)
        df = df.dropna(subset=['keyword', 'domain'])
        
        # Conversion numérique avec debug
        if 'position' in df.columns:
            df['position'] = pd.to_numeric(df['position'], errors='coerce')
            st.info(f"Positions dans {file.name}: min={df['position'].min()}, max={df['position'].max()}, sample={df['position'].head(3).tolist()}")
        
        if 'volume' in df.columns:
            df['volume'] = pd.to_numeric(df['volume'], errors='coerce')
            df['volume'] = df['volume'].fillna(0)  # Remplacer NaN par 0
            st.info(f"Volumes dans {file.name}: min={df['volume'].min()}, max={df['volume'].max()}, sample={df['volume'].head(3).tolist()}")
        
        if 'difficulty' in df.columns:
            df['difficulty'] = pd.to_numeric(df['difficulty'], errors='coerce')
        
        # Filtrer les lignes avec des positions invalides
        if 'position' in df.columns:
            df = df[df['position'] > 0]
        
        # Ajout du nom du fichier
        df['source_file'] = file.name
        
        final_count = len(df)
        st.success(f"Fichier {file.name} traité avec succès : {final_count}/{initial_count} lignes valides")
        
        return df
        
    except Exception as e:
        st.error(f"Erreur lors du chargement de {file.name}: {str(e)}")
        return pd.DataFrame()


def process_files(files, data_source, config):
    """Traite tous les fichiers uploadés"""
    all_dataframes = []
    
    for file in files:
        df = load_file(file, data_source, config)
        if not df.empty:
            all_dataframes.append(df)
    
    if all_dataframes:
        return pd.concat(all_dataframes, ignore_index=True)
    else:
        return pd.DataFrame()


def filter_brand_terms(data, brand_terms):
    """Filtre les mots-clés contenant des termes de marque"""
    if not brand_terms.strip():
        return data
    
    # Nettoyer et séparer les termes
    terms = [term.strip().lower() for term in brand_terms.split(',') if term.strip()]
    
    if not terms:
        return data
    
    # Filtrer les mots-clés contenant ces termes
    def contains_brand_term(keyword):
        if pd.isna(keyword):
            return False
        keyword_lower = str(keyword).lower()
        return any(term in keyword_lower for term in terms)
    
    # Garder seulement les mots-clés qui ne contiennent pas de termes de marque
    filtered_data = data[~data['keyword'].apply(contains_brand_term)]
    
    return filtered_data


def analyze_main_domain(data, main_domain):
    """Analyse le positionnement du domaine principal"""
    if not main_domain:
        return None
    
    # Filtrer les données du domaine principal
    main_data = data[data['domain'] == main_domain].copy()
    
    if main_data.empty:
        return None
    
    # Catégoriser par position
    categories = {
        'Sauvegarde': main_data[(main_data['position'] >= 1) & (main_data['position'] <= 3)],
        'Quick Win': main_data[(main_data['position'] >= 4) & (main_data['position'] <= 5)],
        'Opportunité': main_data[(main_data['position'] >= 6) & (main_data['position'] <= 10)],
        'Potentiel': main_data[(main_data['position'] >= 11) & (main_data['position'] <= 20)],
        'Conquête': main_data[(main_data['position'] >= 21) & (main_data['position'] <= 100)]
    }
    
    # Analyser les mots-clés non positionnés (présents chez les concurrents mais pas chez nous)
    all_keywords = set(data['keyword'].unique())
    main_keywords = set(main_data['keyword'].unique())
    non_positioned = all_keywords - main_keywords
    
    # Récupérer les données des mots-clés non positionnés
    non_positioned_data = data[data['keyword'].isin(non_positioned)].drop_duplicates('keyword')
    
    return {
        'categories': categories,
        'non_positioned': non_positioned_data,
        'main_domain': main_domain
    }


def perform_gap_analysis(data, main_domain, min_competitors, max_position, min_volume, max_difficulty):
    """Effectue l'analyse du gap content - VERSION CORRIGÉE"""
    
    # Filtrage des positions valides
    data = data[(data['position'] <= max_position) & (data['position'] > 0)]
    
    # Filtrage par volume et difficulté
    if min_volume > 0:
        data = data[data['volume'] >= min_volume]
    if max_difficulty < 100:
        data = data[data['difficulty'] <= max_difficulty]
    
    # Groupement par mot-clé
    keyword_analysis = []
    
    for keyword, group in data.groupby('keyword'):
        # Informations du mot-clé
        volume = group['volume'].iloc[0] if not group['volume'].isna().all() else 0
        difficulty = group['difficulty'].iloc[0] if not group['difficulty'].isna().all() else 0
        intent = group['intent'].iloc[0] if 'intent' in group.columns else ''
        
        # Analyse des domaines positionnés
        positioned_domains = group[group['position'] <= max_position]
        unique_domains = positioned_domains['domain'].unique()
        
        # Vérifier si le domaine principal est présent
        main_domain_present = main_domain in unique_domains if main_domain else False
        competitor_count = len(unique_domains) - (1 if main_domain_present else 0)
        
        # Critère de gap content : assez de concurrents positionnés MAIS domaine principal absent
        if competitor_count >= min_competitors and not main_domain_present:
            
            # Trouver la meilleure position globale et l'URL correspondante (IMPORTANT - pas un doublon !)
            global_best_position = positioned_domains['position'].min()
            global_best_url = positioned_domains[positioned_domains['position'] == global_best_position]['url'].iloc[0]
            
            keyword_data = {
                'keyword': keyword,
                'volume': volume,
                'difficulty': difficulty,
                'intent': intent,
                'competitor_count': competitor_count,
                'best_position': global_best_position,  # Meilleure position parmi TOUS les concurrents
                'best_url': global_best_url,  # URL de cette meilleure position
                'total_domains': len(unique_domains)
            }
            
            # CORRECTION 2: Ajouter une colonne pour le domaine principal même s'il n'est pas positionné
            if main_domain:
                main_domain_data = positioned_domains[positioned_domains['domain'] == main_domain]
                if not main_domain_data.empty:
                    # Le domaine principal est positionné (ne devrait pas arriver dans gap content mais sécurité)
                    main_domain_position = main_domain_data['position'].min()
                    main_domain_url = main_domain_data[main_domain_data['position'] == main_domain_position]['url'].iloc[0]
                else:
                    # Le domaine principal n'est pas positionné (cas normal pour gap content)
                    main_domain_position = None
                    main_domain_url = None
                
                # Ajouter les colonnes du domaine principal
                keyword_data[f'{main_domain}_position'] = main_domain_position
                keyword_data[f'{main_domain}_url'] = main_domain_url
            
            # Ajouter les positions et URLs de chaque domaine concurrent
            for domain in unique_domains:
                if domain != main_domain:  # Éviter les doublons avec le domaine principal
                    domain_data = positioned_domains[positioned_domains['domain'] == domain]
                    domain_best_position = domain_data['position'].min()
                    domain_best_url = domain_data[domain_data['position'] == domain_best_position]['url'].iloc[0]
                    
                    keyword_data[f'{domain}_position'] = domain_best_position
                    keyword_data[f'{domain}_url'] = domain_best_url
            
            keyword_analysis.append(keyword_data)
    
    gap_content_df = pd.DataFrame(keyword_analysis)
    
    # Création du rapport par domaine
    domain_reports = {}
    for domain in data['domain'].unique():
        if domain and domain != main_domain:
            domain_data = data[data['domain'] == domain].copy()
            domain_data = domain_data.sort_values(['position', 'volume'], ascending=[True, False])
            domain_reports[domain] = domain_data[['keyword', 'volume', 'position', 'url']].reset_index(drop=True)
    
    return {
        'gap_content': gap_content_df,
        'domain_reports': domain_reports,
        'main_domain': main_domain,
        'all_data': data
    }


def generate_excel_report(analysis, main_domain, main_domain_analysis=None):
    """Génère le rapport Excel avec mise en forme - VERSION CORRIGÉE AVEC ESTHÉTIQUE"""
    
    output = io.BytesIO()
    workbook = Workbook()
    
    # Suppression de la feuille par défaut
    workbook.remove(workbook.active)
    
    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    main_domain_fill = PatternFill(start_color="FFFFB7", end_color="FFFFB7", fill_type="solid")  # Jaune pour domaine principal
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Couleurs pour les intentions de recherche
    intent_colors = {
        'informational': 'ADE8F4',
        'commercial': 'FFFFB7', 
        'transactional': 'D8F3DC',
        'navigational': 'B79CED'
    }
    
    def hex_to_rgb(hex_color):
        """Convertit une couleur hex en RGB"""
        hex_color = hex_color.lstrip('#')
        return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    
    def rgb_to_hex(rgb):
        """Convertit RGB en hex"""
        return f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"
    
    def interpolate_color(color1_hex, color2_hex, factor):
        """Interpole entre deux couleurs (factor entre 0 et 1)"""
        rgb1 = hex_to_rgb(color1_hex)
        rgb2 = hex_to_rgb(color2_hex)
        
        interpolated = tuple(
            int(rgb1[i] + factor * (rgb2[i] - rgb1[i]))
            for i in range(3)
        )
        return rgb_to_hex(interpolated)
    
    def get_gradient_color(value, min_val, max_val, min_color, max_color):
        """Calcule la couleur dans un dégradé selon la valeur"""
        if max_val == min_val:
            return min_color
        factor = (value - min_val) / (max_val - min_val)
        return interpolate_color(min_color, max_color, factor)
    
    # 1. Onglet Gap Content Analysis - VERSION CORRIGÉE
    if not analysis['gap_content'].empty:
        ws_gap = workbook.create_sheet("Gap Content Analysis")
        gap_df = analysis['gap_content'].copy()
        
        # CORRECTION 3: Réorganiser les colonnes en mettant le domaine principal EN PREMIER
        base_cols = ['keyword', 'volume', 'difficulty', 'intent', 'competitor_count', 'best_position', 'best_url']
        
        # Identifier tous les domaines
        position_columns = [col for col in gap_df.columns if col.endswith('_position')]
        url_columns = [col for col in gap_df.columns if col.endswith('_url')]
        
        # Extraire les noms de domaines
        all_domains = list(set([col.replace('_position', '').replace('_url', '') for col in position_columns + url_columns]))
        
        # CORRECTION: Mettre le domaine principal en premier dans l'ordre
        ordered_domains = []
        if main_domain and main_domain in all_domains:
            ordered_domains.append(main_domain)
        
        # Ajouter les autres domaines triés
        for domain in sorted(all_domains):
            if domain != main_domain and domain not in ordered_domains:
                ordered_domains.append(domain)
        
        # Construire l'ordre des colonnes : base + domaine principal + autres domaines (positions puis URLs)
        ordered_position_cols = []
        ordered_url_cols = []
        
        for domain in ordered_domains:
            if f'{domain}_position' in gap_df.columns:
                ordered_position_cols.append(f'{domain}_position')
            if f'{domain}_url' in gap_df.columns:
                ordered_url_cols.append(f'{domain}_url')
        
        all_cols = base_cols + ordered_position_cols + ordered_url_cols
        gap_df_display = gap_df[all_cols].copy()
        
        # Renommer les colonnes pour l'affichage
        column_mapping = {
            'keyword': 'Mot-clé',
            'volume': 'Volume de recherche',
            'difficulty': 'Difficulté concurrentielle',
            'intent': 'Intention de recherche',
            'competitor_count': 'Nombre de concurrents positionnés',
            'best_position': 'Meilleure position globale',
            'best_url': 'URL de la meilleure position'
        }
        
        # Ajouter les noms de domaines dans le mapping
        for domain in ordered_domains:
            domain_name = domain.replace('_', '.')
            if f'{domain}_position' in gap_df.columns:
                if domain == main_domain:
                    column_mapping[f'{domain}_position'] = f'🏠 {domain_name} (Position)'
                else:
                    column_mapping[f'{domain}_position'] = f'{domain_name} (Position)'
            if f'{domain}_url' in gap_df.columns:
                if domain == main_domain:
                    column_mapping[f'{domain}_url'] = f'🏠 {domain_name} (URL)'
                else:
                    column_mapping[f'{domain}_url'] = f'{domain_name} (URL)'
        
        gap_df_display = gap_df_display.rename(columns=column_mapping)
        
        # Écriture des données
        for r in dataframe_to_rows(gap_df_display, index=False, header=True):
            ws_gap.append(r)
        
        # Mise en forme des en-têtes
        for col_idx, cell in enumerate(ws_gap[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            
            # Mise en forme spéciale pour les colonnes du domaine principal
            if main_domain and (f'🏠 {main_domain}' in str(cell.value)):
                cell.fill = main_domain_fill
                cell.font = Font(color="000000", bold=True)  # Texte noir sur fond jaune
        
        # NOUVELLES MISES EN FORME ESTHÉTIQUES
        
        # 1. Préparation des données pour les dégradés
        if len(gap_df_display) > 0:
            # Trouver les colonnes par leur nom français
            col_indices = {}
            for col_idx, cell in enumerate(ws_gap[1], 1):
                col_name = str(cell.value)
                if col_name == 'Nombre de concurrents positionnés':
                    col_indices['competitors'] = col_idx
                elif col_name == 'Intention de recherche':
                    col_indices['intent'] = col_idx
                elif col_name == 'Difficulté concurrentielle':
                    col_indices['difficulty'] = col_idx
                elif col_name == 'Volume de recherche':
                    col_indices['volume'] = col_idx
                elif main_domain and f'🏠 {main_domain}' in col_name:
                    if 'main_domain_cols' not in col_indices:
                        col_indices['main_domain_cols'] = []
                    col_indices['main_domain_cols'].append(col_idx)
            
            # Calculer les valeurs min/max pour les dégradés
            if 'competitors' in col_indices:
                competitor_values = [ws_gap.cell(row=i, column=col_indices['competitors']).value 
                                   for i in range(2, ws_gap.max_row + 1)]
                competitor_values = [v for v in competitor_values if v is not None]
                if competitor_values:
                    min_competitors = min(competitor_values)
                    max_competitors = max(competitor_values)
            
            if 'difficulty' in col_indices:
                difficulty_values = [ws_gap.cell(row=i, column=col_indices['difficulty']).value 
                                   for i in range(2, ws_gap.max_row + 1)]
                difficulty_values = [v for v in difficulty_values if v is not None and str(v).replace('.','').isdigit()]
                if difficulty_values:
                    min_difficulty = min(difficulty_values)
                    max_difficulty = max(difficulty_values)
                    median_difficulty = sorted(difficulty_values)[len(difficulty_values)//2]
            
            if 'volume' in col_indices:
                volume_values = [ws_gap.cell(row=i, column=col_indices['volume']).value 
                               for i in range(2, ws_gap.max_row + 1)]
                volume_values = [v for v in volume_values if v is not None and str(v).replace('.','').isdigit()]
                if volume_values:
                    min_volume = min(volume_values)
                    max_volume = max(volume_values)
            
            # 2. Application des couleurs ligne par ligne
            for row_idx in range(2, ws_gap.max_row + 1):
                
                # Coloration "Nombre de concurrents positionnés"
                if 'competitors' in col_indices and competitor_values:
                    cell = ws_gap.cell(row=row_idx, column=col_indices['competitors'])
                    if cell.value is not None:
                        color = get_gradient_color(cell.value, min_competitors, max_competitors, 
                                                 'F6CACC', 'DD2C2F')
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                
                # Coloration "Intention de recherche"
                if 'intent' in col_indices:
                    cell = ws_gap.cell(row=row_idx, column=col_indices['intent'])
                    if cell.value is not None:
                        intent_value = str(cell.value).lower().split(',')[0].strip()  # Première valeur avant la virgule
                        if intent_value in intent_colors:
                            color = intent_colors[intent_value]
                            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                
                # Coloration "Difficulté concurrentielle"
                if 'difficulty' in col_indices and difficulty_values:
                    cell = ws_gap.cell(row=row_idx, column=col_indices['difficulty'])
                    if cell.value is not None and str(cell.value).replace('.','').isdigit():
                        difficulty_val = float(cell.value)
                        if difficulty_val <= median_difficulty:
                            # Du minimum (vert) vers la médiane (jaune)
                            factor = (difficulty_val - min_difficulty) / (median_difficulty - min_difficulty) if median_difficulty != min_difficulty else 0
                            color = interpolate_color('D8F3DC', 'FFFFB7', factor)
                        else:
                            # De la médiane (jaune) vers le maximum (rouge)
                            factor = (difficulty_val - median_difficulty) / (max_difficulty - median_difficulty) if max_difficulty != median_difficulty else 0
                            color = interpolate_color('FFFFB7', 'DD2C2F', factor)
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                
                # Coloration "Volume de recherche"
                if 'volume' in col_indices and volume_values:
                    cell = ws_gap.cell(row=row_idx, column=col_indices['volume'])
                    if cell.value is not None and str(cell.value).replace('.','').isdigit():
                        # Dégradé du blanc au vert foncé basé sur le volume
                        factor = (float(cell.value) - min_volume) / (max_volume - min_volume) if max_volume != min_volume else 0
                        # Dégradé du blanc (FFFFFF) vers le vert (52B788)
                        color = interpolate_color('FFFFFF', '52B788', factor)
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                
                # Coloration des colonnes du domaine principal
                if 'main_domain_cols' in col_indices:
                    for col_idx in col_indices['main_domain_cols']:
                        cell = ws_gap.cell(row=row_idx, column=col_idx)
                        cell.fill = PatternFill(start_color="FFFFB7", end_color="FFFFB7", fill_type="solid")  # Jaune
        
        # Mise en forme des données - colonnes du domaine principal en jaune (déjà fait ci-dessus)
        
        # Ajustement des largeurs de colonnes
        for col_num in range(1, len(gap_df_display.columns) + 1):
            column_letter = ws_gap.cell(row=1, column=col_num).column_letter
            max_length = 0
            for row_num in range(1, ws_gap.max_row + 1):
                cell = ws_gap.cell(row=row_num, column=col_num)
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_gap.column_dimensions[column_letter].width = adjusted_width
    
    # 2. Onglet récapitulatif du domaine principal
    if main_domain_analysis:
        ws_main = workbook.create_sheet(f"Analyse {main_domain}")
        
        # Titre principal
        ws_main.append([f"Analyse de positionnement - {main_domain}"])
        ws_main['A1'].font = Font(size=18, bold=True)
        ws_main.merge_cells('A1:D1')
        ws_main.append([])
        
        current_row = 3
        
        # Pour chaque catégorie
        categories_order = ['Sauvegarde', 'Quick Win', 'Opportunité', 'Potentiel', 'Conquête']
        position_ranges = {
            'Sauvegarde': '1-3',
            'Quick Win': '4-5', 
            'Opportunité': '6-10',
            'Potentiel': '11-20',
            'Conquête': '21-100'
        }
        
        for category in categories_order:
            cat_data = main_domain_analysis['categories'][category]
            
            if not cat_data.empty:
                # Titre de catégorie
                ws_main.append([f"{category} (Positions {position_ranges[category]}) - {len(cat_data)} mots-clés"])
                ws_main[f'A{current_row}'].font = Font(size=14, bold=True)
                current_row += 1
                
                # En-têtes
                headers = ['Mot-clé', 'Volume de recherche', 'Position', 'URL']
                ws_main.append(headers)
                for cell in ws_main[current_row]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment
                current_row += 1
                
                # Données triées par position puis volume
                sorted_data = cat_data.sort_values(['position', 'volume'], ascending=[True, False])
                for _, row in sorted_data.iterrows():
                    ws_main.append([row['keyword'], row['volume'], row['position'], row['url']])
                    current_row += 1
                
                # Mise en forme conditionnelle des volumes
                if len(sorted_data) > 0:
                    min_volume = sorted_data['volume'].min()
                    max_volume = sorted_data['volume'].max()
                    
                    start_row = current_row - len(sorted_data)
                    for row_num in range(start_row, current_row):
                        cell = ws_main[f'B{row_num}']
                        volume = cell.value
                        if volume and max_volume > min_volume:
                            intensity = (volume - min_volume) / (max_volume - min_volume)
                            green_value = int(255 - (intensity * 100))
                            color = f"FF{green_value:02X}FF{green_value:02X}"
                            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                
                ws_main.append([])  # Ligne vide
                current_row += 1
        
        # Section Non positionné
        non_pos_data = main_domain_analysis['non_positioned']
        if not non_pos_data.empty:
            ws_main.append([f"Non positionné - {len(non_pos_data)} mots-clés"])
            ws_main[f'A{current_row}'].font = Font(size=14, bold=True)
            current_row += 1
            
            # En-têtes
            headers = ['Mot-clé', 'Volume de recherche', 'Difficulté', 'Intention']
            ws_main.append(headers)
            for cell in ws_main[current_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            current_row += 1
            
            # Données triées par volume
            sorted_data = non_pos_data.sort_values('volume', ascending=False)
            for _, row in sorted_data.iterrows():
                ws_main.append([row['keyword'], row['volume'], row.get('difficulty', ''), row.get('intent', '')])
                current_row += 1
        
        # Ajustement des largeurs
        for col_num in range(1, 5):
            column_letter = ws_main.cell(row=3, column=col_num).column_letter
            max_length = 0
            for row_num in range(1, ws_main.max_row + 1):
                cell = ws_main.cell(row=row_num, column=col_num)
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_main.column_dimensions[column_letter].width = adjusted_width
    
    # 3. Onglets pour chaque domaine concurrent
    for domain, domain_data in analysis['domain_reports'].items():
        if not domain_data.empty:
            # Créer un nom d'onglet valide (max 31 caractères)
            sheet_name = domain.replace('www.', '').replace('.com', '').replace('.fr', '')[:31]
            ws_domain = workbook.create_sheet(sheet_name)
            
            # Ligne de titre avec le nom du domaine
            ws_domain.append([domain])
            ws_domain['A1'].font = Font(size=18, bold=True)
            ws_domain.merge_cells('A1:D1')
            
            # Ligne vide
            ws_domain.append([])
            
            # En-têtes
            headers = ['Mot-clé', 'Volume de recherche', 'Position', 'URL']
            ws_domain.append(headers)
            
            # Mise en forme des en-têtes
            for cell in ws_domain[3]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            
            # Données triées par position puis volume
            sorted_data = domain_data.sort_values(['position', 'volume'], ascending=[True, False])
            
            # Écriture des données
            for _, row in sorted_data.iterrows():
                ws_domain.append([row['keyword'], row['volume'], row['position'], row['url']])
            
            # Mise en forme conditionnelle pour les volumes
            if len(sorted_data) > 0:
                min_volume = sorted_data['volume'].min()
                max_volume = sorted_data['volume'].max()
                
                # Coloration conditionnelle des volumes (colonne B)
                for row_num in range(4, len(sorted_data) + 4):
                    cell = ws_domain[f'B{row_num}']
                    volume = cell.value
                    if volume and max_volume > min_volume:
                        # Gradient du blanc au vert foncé
                        intensity = (volume - min_volume) / (max_volume - min_volume)
                        green_value = int(255 - (intensity * 100))  # De 255 (blanc) à 155 (vert clair)
                        color = f"FF{green_value:02X}FF{green_value:02X}"
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            # Ajustement des largeurs
            for col_num in range(1, 5):  # 4 colonnes : Mot-clé, Volume, Position, URL
                column_letter = ws_domain.cell(row=3, column=col_num).column_letter
                max_length = 0
                for row_num in range(1, ws_domain.max_row + 1):
                    cell = ws_domain.cell(row=row_num, column=col_num)
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_domain.column_dimensions[column_letter].width = adjusted_width
    
    # Sauvegarde
    workbook.save(output)
    output.seek(0)
    
    return output.getvalue()


if __name__ == "__main__":
    main()
